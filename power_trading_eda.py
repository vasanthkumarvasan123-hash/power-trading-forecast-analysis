# =============================================================================
# IEX WEATHER DATASET - COMPLETE ANALYSIS
# Data Cleaning + Preprocessing + 4 Moments + All Graphs
# SPYDER VERSION: All plots show in Spyder Plots panel + saved as PNG
# =============================================================================

import pandas as pd
import numpy as np

# ── Spyder inline backend ── #
import matplotlib
matplotlib.use('inline')          # Showing the Spyder Plots panel
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from scipy import stats
import warnings
import os
warnings.filterwarnings('ignore')

# Spyder-la plt.show() work aganum-na ipdi set pannu
plt.rcParams['figure.dpi']     = 100
plt.rcParams['figure.figsize'] = (7, 4.5)

# =============================================================================
# STEP 1: LOAD DATA
# =============================================================================
FILE_PATH = r"C:\Users\welcome\Downloads\Dataset\New folder\IEX_Weather_final.xlsx"

print("=" * 70)
print("  IEX WEATHER DATASET - COMPLETE STATISTICAL ANALYSIS")
print("=" * 70)
print("\nLoading data...")

df = pd.read_excel(FILE_PATH)

print(f"Data Loaded Successfully!")
print(f"   Shape       : {df.shape[0]} rows x {df.shape[1]} columns")
print(f"   Memory Usage: {df.memory_usage(deep=True).sum() / 1e6:.2f} MB")

# =============================================================================
# STEP 2: DATA CLEANING & PREPROCESSING
# =============================================================================
print("\n" + "=" * 70)
print("  STEP 2: DATA CLEANING & PREPROCESSING")
print("=" * 70)

null_total = df.isnull().sum().sum()
print(f"\nTotal missing values: {null_total}")

dup_count = df.duplicated().sum()
if dup_count > 0:
    df = df.drop_duplicates()
    print(f"   {dup_count} duplicate rows removed!")
else:
    print("   No duplicates found!")

if 'Date' in df.columns:
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
if 'time' in df.columns:
    df['time'] = pd.to_datetime(df['time'], errors='coerce')

skip_cols    = ['Date', 'time', 'Time Block']
numeric_cols = [
    col for col in df.columns
    if col not in skip_cols and pd.api.types.is_numeric_dtype(df[col])
]

print(f"\nTotal Numeric Columns for Analysis: {len(numeric_cols)}")

outlier_info = {}
for col in numeric_cols:
    Q1    = df[col].quantile(0.25)
    Q3    = df[col].quantile(0.75)
    IQR   = Q3 - Q1
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR
    n_out = ((df[col] < lower) | (df[col] > upper)).sum()
    outlier_info[col] = {'lower': lower, 'upper': upper, 'count': n_out}

for col in numeric_cols:
    if df[col].isnull().sum() > 0:
        df[col].fillna(df[col].median(), inplace=True)

print("DATA CLEANING COMPLETE!")

# =============================================================================
# STEP 3: COMPUTE 4 STATISTICAL MOMENTS
# =============================================================================
def compute_4_moments(series, col_name):
    s        = series.dropna()
    mean     = s.mean()
    variance = s.var(ddof=1)
    std_dev  = s.std(ddof=1)
    skewness = stats.skew(s)
    kurt     = stats.kurtosis(s)

    if   skewness >  0.5: skew_interp = "Right Skewed (+)"
    elif skewness < -0.5: skew_interp = "Left Skewed  (-)"
    else:                  skew_interp = "Approximately Symmetric"

    if   kurt > 0: kurt_interp = "Leptokurtic (Heavy tails)"
    elif kurt < 0: kurt_interp = "Platykurtic  (Light tails)"
    else:           kurt_interp = "Mesokurtic   (Normal tails)"

    return {
        'Column'     : col_name,
        'Count'      : len(s),
        'Mean'       : round(mean, 4),
        'Variance'   : round(variance, 4),
        'Std_Dev'    : round(std_dev, 4),
        'Skewness'   : round(skewness, 4),
        'Kurtosis'   : round(kurt, 4),
        'Skew_Interp': skew_interp,
        'Kurt_Interp': kurt_interp
    }

all_moments = [compute_4_moments(df[col], col) for col in numeric_cols]
moments_df  = pd.DataFrame(all_moments)
moments_df.to_csv("IEX_4Moments_Results.csv", index=False)
print("4 Moments saved: IEX_4Moments_Results.csv")

# =============================================================================
# STEP 4: PLOT FUNCTIONS  (savefig + plt.show — both work)
# =============================================================================

SCATTER_REF = 'MCP (Rs/MWh) *' if 'MCP (Rs/MWh) *' in df.columns else numeric_cols[0]
os.makedirs("plots", exist_ok=True)

def safe_name(col, col_num):
    clean = (col[:30]
             .replace('/', '_').replace(' ', '_')
             .replace('(', '').replace(')', '')
             .replace('*', '').replace('\u00b0', '')
             .replace('%', 'pct').replace('\u00b2', 'sq'))
    return f"row{col_num+3:02d}_{clean}"

def show_and_save(fname):
    """Save PNG and show in Spyder Plots panel."""
    plt.tight_layout()
    plt.savefig(fname, dpi=100, bbox_inches='tight')
    plt.show()          # <-- Spyder Plots panel-la kaatum
    plt.close()

# ------------------------------------------------------------------
# HISTOGRAM
# ------------------------------------------------------------------
def plot_histogram(series, col, col_num, moments_row):
    try:
        fig, ax = plt.subplots(figsize=(7, 4.5))
        ax.hist(series, bins=50, color='steelblue', edgecolor='white',
                alpha=0.75, density=True)
        kde_x = np.linspace(series.min(), series.max(), 300)
        kde   = stats.gaussian_kde(series)
        ax.plot(kde_x, kde(kde_x), color='crimson', linewidth=2, label='KDE')
        ax.axvline(series.mean(),   color='orange', linestyle='--',
                   linewidth=1.5, label=f'Mean={series.mean():.2f}')
        ax.axvline(series.median(), color='green',  linestyle=':',
                   linewidth=1.5, label=f'Median={series.median():.2f}')
        ax.set_title(f'Histogram: {col[:50]}', fontsize=9, fontweight='bold')
        ax.set_xlabel(col[:50], fontsize=8)
        ax.set_ylabel('Density', fontsize=8)
        ax.legend(fontsize=7)
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.3)
        skewness = moments_row['Skewness']
        ax.text(0.97, 0.95,
                f'Skew: {skewness:.3f}\n{moments_row["Skew_Interp"]}',
                transform=ax.transAxes, ha='right', va='top', fontsize=7,
                bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.8))
        fname = f"plots/{safe_name(col, col_num)}_hist.png"
        show_and_save(fname)
        return fname
    except Exception as e:
        plt.close()
        print(f"  [WARN] Histogram failed for {col}: {e}")
        return None

# ------------------------------------------------------------------
# BOXPLOT
# ------------------------------------------------------------------
def plot_boxplot(series, col, col_num, moments_row):
    try:
        fig, ax = plt.subplots(figsize=(7, 4.5))
        ax.boxplot(series, vert=True, patch_artist=True,
                   boxprops=dict(facecolor='lightblue', color='navy'),
                   medianprops=dict(color='red', linewidth=2),
                   whiskerprops=dict(color='navy'),
                   capprops=dict(color='navy'),
                   flierprops=dict(marker='o', color='gray',
                                   alpha=0.3, markersize=3))
        Q1    = series.quantile(0.25)
        Q3    = series.quantile(0.75)
        IQR   = Q3 - Q1
        n_out = ((series < Q1 - 1.5*IQR) | (series > Q3 + 1.5*IQR)).sum()
        ax.set_title(f'Boxplot: {col[:50]}', fontsize=9, fontweight='bold')
        ax.set_ylabel(col[:50], fontsize=8)
        ax.set_xticklabels([col[:25]], fontsize=7)
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.3, axis='y')
        ax.text(0.97, 0.97,
                f'Q1={Q1:.2f}\nQ3={Q3:.2f}\nIQR={IQR:.2f}\nOutliers={n_out}',
                transform=ax.transAxes, ha='right', va='top', fontsize=7,
                bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.8))
        fname = f"plots/{safe_name(col, col_num)}_box.png"
        show_and_save(fname)
        return fname
    except Exception as e:
        plt.close()
        print(f"  [WARN] Boxplot failed for {col}: {e}")
        return None

# ------------------------------------------------------------------
# SCATTER
# ------------------------------------------------------------------
def plot_scatter(df, series, col, col_num, ref_col, moments_row):
    try:
        fig, ax = plt.subplots(figsize=(7, 4.5))
        if col != ref_col and ref_col in df.columns:
            ref_series = df[ref_col].dropna()
            common_idx = series.index.intersection(ref_series.index)
            x_vals = ref_series.loc[common_idx]
            y_vals = series.loc[common_idx]
            if len(x_vals) > 5000:
                sample_idx = np.random.choice(len(x_vals), 5000, replace=False)
                x_plot = x_vals.iloc[sample_idx]
                y_plot = y_vals.iloc[sample_idx]
            else:
                x_plot, y_plot = x_vals, y_vals
            ax.scatter(x_plot, y_plot, alpha=0.3, color='darkorchid',
                       s=8, edgecolors='none')
            if len(x_plot) > 2:
                slope, intercept, r_val, _, _ = stats.linregress(x_plot, y_plot)
                x_line = np.linspace(x_plot.min(), x_plot.max(), 100)
                ax.plot(x_line, slope * x_line + intercept, color='red',
                        linewidth=2, label=f'R={r_val:.3f}')
                ax.legend(fontsize=8)
            ax.set_xlabel(ref_col[:50], fontsize=8)
            ax.set_ylabel(col[:50], fontsize=8)
        else:
            idx_vals = np.arange(min(5000, len(series)))
            ax.scatter(idx_vals, series.iloc[:5000], alpha=0.3, color='teal', s=5)
            ax.set_xlabel('Index', fontsize=8)
            ax.set_ylabel(col[:50], fontsize=8)
        ax.set_title(f'Scatter: {col[:40]} vs {ref_col[:20]}',
                     fontsize=9, fontweight='bold')
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.3)
        fname = f"plots/{safe_name(col, col_num)}_scatter.png"
        show_and_save(fname)
        return fname
    except Exception as e:
        plt.close()
        print(f"  [WARN] Scatter failed for {col}: {e}")
        return None

# ------------------------------------------------------------------
# MULTIVARIATE
# ------------------------------------------------------------------
def get_col_type(col):
    if col == 'Hour':                            return 'hour'
    if col in ['Purchase Bid (MW)', 'Sell Bid (MW)', 'MCV (MW)',
               'Final Scheduled Volume (MW)', 'MCP (Rs/MWh) *']:
        return 'market'
    if 'temperature'         in col:             return 'temperature'
    if 'relative_humidity'   in col:             return 'humidity'
    if 'cloud_cover'         in col:             return 'cloud'
    if 'wind_speed'          in col:             return 'wind'
    if 'shortwave_radiation' in col:             return 'radiation'
    return 'other'

def plot_multivariate(df, series, col, col_num, moments_row):
    fig, ax = plt.subplots(figsize=(7, 4.5))
    col_type = get_col_type(col)
    mcp_col  = 'MCP (Rs/MWh) *' if 'MCP (Rs/MWh) *' in df.columns else None

    def sample_data(x, y, z=None, n=3000):
        idx = x.dropna().index.intersection(y.dropna().index)
        if z is not None:
            idx = idx.intersection(z.dropna().index)
        if len(idx) > n:
            idx = idx[np.random.choice(len(idx), n, replace=False)]
        if z is not None:
            return x.loc[idx], y.loc[idx], z.loc[idx]
        return x.loc[idx], y.loc[idx]

    def scatter3(x_s, y_s, c_s, xlabel, ylabel, clabel, title):
        sc = ax.scatter(x_s, y_s, c=c_s, cmap='plasma',
                        alpha=0.5, s=10, edgecolors='none')
        cb = plt.colorbar(sc, ax=ax, pad=0.01)
        cb.set_label(clabel[:25], fontsize=6)
        cb.ax.tick_params(labelsize=6)
        ax.set_xlabel(xlabel[:40], fontsize=8)
        ax.set_ylabel(ylabel[:40], fontsize=8)
        ax.set_title(title[:50], fontsize=9, fontweight='bold')
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.3)

    try:
        if col_type == 'hour':
            if mcp_col:
                temp_df = pd.DataFrame(
                    {'hour': df['Hour'], 'mcp': df[mcp_col]}).dropna()
                pivot = temp_df.groupby('hour')['mcp'].mean()
                hours = pivot.index.values
                vals  = pivot.values.reshape(1, -1)
                im = ax.imshow(vals, aspect='auto', cmap='YlOrRd',
                               extent=[hours.min()-0.5, hours.max()+0.5, 0, 1])
                plt.colorbar(im, ax=ax, label='Avg MCP (Rs/MWh)',
                             pad=0.01).ax.tick_params(labelsize=6)
                ax.set_xlabel('Hour of Day', fontsize=8)
                ax.set_yticks([])
                ax.set_title('Heatmap: Hour vs Avg MCP', fontsize=9,
                             fontweight='bold')
                ax.tick_params(labelsize=7)
            else:
                ax.hist(series, bins=24, color='steelblue', edgecolor='white')
                ax.set_title('Hour Distribution', fontsize=9, fontweight='bold')

        elif col_type == 'market':
            if mcp_col and mcp_col != col and 'Hour' in df.columns:
                x_s, y_s, c_s = sample_data(series, df[mcp_col], df['Hour'])
                scatter3(x_s, y_s, c_s,
                         col[:30], 'MCP (Rs/MWh)', 'Hour',
                         f'{col[:20]} vs MCP (color=Hour)')
            else:
                idx_vals = np.arange(min(3000, len(series)))
                ax.scatter(idx_vals, series.iloc[:3000],
                           alpha=0.4, color='darkorchid', s=6)
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        elif col_type == 'temperature':
            loc_prefix = col.split('_')[0]
            cloud_col  = next((c for c in df.columns
                               if c.startswith(loc_prefix) and 'cloud_cover' in c), None)
            if mcp_col and cloud_col:
                x_s, y_s, c_s = sample_data(series, df[mcp_col], df[cloud_col])
                scatter3(x_s, y_s, c_s,
                         'Temperature (°C)', 'MCP (Rs/MWh)', 'Cloud Cover (%)',
                         'Temp vs MCP (color=Cloud)')
            else:
                idx_vals = np.arange(min(3000, len(series)))
                ax.scatter(idx_vals, series.iloc[:3000],
                           color='tomato', s=6, alpha=0.4)
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        elif col_type == 'humidity':
            loc_prefix = col.split('_')[0]
            temp_col   = next((c for c in df.columns
                               if c.startswith(loc_prefix) and 'temperature' in c), None)
            if mcp_col and temp_col:
                x_s, y_s, c_s = sample_data(series, df[temp_col], df[mcp_col])
                scatter3(x_s, y_s, c_s,
                         'Humidity (%)', 'Temperature (°C)', 'MCP (Rs/MWh)',
                         'Humidity vs Temp (color=MCP)')
            else:
                ax.hist(series, bins=30, color='skyblue', edgecolor='white')
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        elif col_type == 'cloud':
            loc_prefix = col.split('_')[0]
            rad_col    = next((c for c in df.columns
                               if c.startswith(loc_prefix) and 'shortwave_radiation' in c), None)
            if mcp_col and rad_col:
                x_s, y_s, c_s = sample_data(series, df[rad_col], df[mcp_col])
                scatter3(x_s, y_s, c_s,
                         'Cloud Cover (%)', 'Solar Radiation (W/m²)', 'MCP (Rs/MWh)',
                         'Cloud vs Radiation (color=MCP)')
            else:
                ax.hist(series, bins=30, color='lightgray', edgecolor='white')
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        elif col_type == 'wind':
            loc_prefix = col.split('_')[0]
            rad_col    = next((c for c in df.columns
                               if c.startswith(loc_prefix) and 'shortwave_radiation' in c), None)
            if mcp_col and rad_col:
                x_s, y_s, c_s = sample_data(series, df[mcp_col], df[rad_col])
                scatter3(x_s, y_s, c_s,
                         'Wind Speed (km/h)', 'MCP (Rs/MWh)', 'Radiation (W/m²)',
                         'Wind vs MCP (color=Radiation)')
            else:
                ax.hist(series, bins=30, color='lightgreen', edgecolor='white')
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        elif col_type == 'radiation':
            loc_prefix = col.split('_')[0]
            cloud_col  = next((c for c in df.columns
                               if c.startswith(loc_prefix) and 'cloud_cover' in c), None)
            if mcp_col and cloud_col:
                x_s, y_s, c_s = sample_data(series, df[mcp_col], df[cloud_col])
                scatter3(x_s, y_s, c_s,
                         'Radiation (W/m²)', 'MCP (Rs/MWh)', 'Cloud Cover (%)',
                         'Radiation vs MCP (color=Cloud)')
            else:
                ax.hist(series, bins=30, color='gold', edgecolor='white')
                ax.set_title(col[:45], fontsize=9, fontweight='bold')

        else:   # violin by time of day
            if 'Hour' in df.columns:
                hour_col = df['Hour']
                def hour_group(h):
                    if   h in range(6,  12): return 'Morning\n(6-12)'
                    elif h in range(12, 18): return 'Afternoon\n(12-18)'
                    elif h in range(18, 24): return 'Evening\n(18-24)'
                    else:                    return 'Night\n(0-6)'
                temp_df2 = pd.DataFrame(
                    {'value': series, 'group': hour_col.map(hour_group)}).dropna()
                groups  = ['Night\n(0-6)', 'Morning\n(6-12)',
                           'Afternoon\n(12-18)', 'Evening\n(18-24)']
                gdata   = [temp_df2[temp_df2['group'] == g]['value'].values
                           for g in groups if len(temp_df2[temp_df2['group'] == g]) > 0]
                glabels = [g for g in groups
                           if len(temp_df2[temp_df2['group'] == g]) > 0]
                if len(gdata) > 1:
                    parts = ax.violinplot(gdata, positions=range(len(gdata)),
                                         showmeans=True, showmedians=True)
                    for pc in parts['bodies']:
                        pc.set_facecolor('lightcoral')
                        pc.set_alpha(0.7)
                    ax.set_xticks(range(len(glabels)))
                    ax.set_xticklabels(glabels, fontsize=7)
                ax.set_title(f'Violin by Time: {col[:35]}', fontsize=9,
                             fontweight='bold')
                ax.set_ylabel(col[:40], fontsize=8)
                ax.grid(True, alpha=0.3, axis='y')

    except Exception as e:
        ax.text(0.5, 0.5, f'Plot error:\n{str(e)[:80]}',
                transform=ax.transAxes, ha='center', va='center',
                fontsize=7, color='red')
        ax.set_title(col[:45], fontsize=9)

    fname = f"plots/{safe_name(col, col_num)}_multivariate.png"
    show_and_save(fname)
    return fname

# =============================================================================
# STEP 5: GENERATE ALL PLOTS
# =============================================================================
print("\nGenerating all plots — each will appear in Spyder Plots panel...")
print("(76 columns x 4 plots = 304 plots total)\n")

total        = len(numeric_cols)
moments_dict = moments_df.set_index('Column').to_dict('index')
plot_files   = {}

for i, col in enumerate(numeric_cols, start=1):
    m_row  = moments_dict[col]
    series = df[col].dropna()

    print(f"  Plotting [{i}/{total}]: {col[:55]}")

    hist_f    = plot_histogram(series, col, i, m_row)
    box_f     = plot_boxplot(series, col, i, m_row)
    scatter_f = plot_scatter(df, series, col, i, SCATTER_REF, m_row)
    multi_f   = plot_multivariate(df, series, col, i, m_row)

    plot_files[col] = {
        'hist'        : hist_f,
        'box'         : box_f,
        'scatter'     : scatter_f,
        'multivariate': multi_f,
        'excel_row'   : i + 3
    }

print(f"\nAll {total*4} plots shown in Spyder + saved in 'plots/' folder!")

# =============================================================================
# STEP 6: EMBED PLOTS INTO EXCEL
# =============================================================================
EXCEL_INPUT  = 'IEX_EDA_Business_Insights_NEW.xlsx'
EXCEL_OUTPUT = 'IEX_EDA_Business_Insights_WITH_PLOTS.xlsx'

if not os.path.exists(EXCEL_INPUT):
    print(f"\n[WARN] '{EXCEL_INPUT}' not found — skipping Excel embedding.")
    print("       Place the file next to this script and re-run to embed plots.")
else:
    print("\nEmbedding plots into Excel...")

    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    wb = load_workbook(EXCEL_INPUT)
    ws = wb['Business Insights']

    COL_HIST    = 14   # N
    COL_BOX     = 17   # Q
    COL_SCATTER = 20   # T
    COL_VIOLIN  = 23   # W
    IMG_HEIGHT_PX = 110
    IMG_WIDTH_PX  = 150

    for col_name, files in plot_files.items():
        excel_row = files['excel_row']
        ws.row_dimensions[excel_row].height = 85

        def embed(png_path, col_idx):
            if png_path is None or not os.path.exists(png_path):
                return
            try:
                img        = XLImage(png_path)
                img.width  = IMG_WIDTH_PX
                img.height = IMG_HEIGHT_PX
                cell_addr  = ws.cell(row=excel_row, column=col_idx).coordinate
                ws.add_image(img, cell_addr)
            except Exception as e:
                print(f"  [WARN] Could not embed {png_path}: {e}")

        embed(files['hist'],         COL_HIST)
        embed(files['box'],          COL_BOX)
        embed(files['scatter'],      COL_SCATTER)
        embed(files['multivariate'], COL_VIOLIN)

    for col_idx in [COL_HIST, COL_BOX, COL_SCATTER, COL_VIOLIN]:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    wb.save(EXCEL_OUTPUT)
    print(f"Excel with embedded plots saved: {EXCEL_OUTPUT}")

# =============================================================================
# STEP 7: SUMMARY
# =============================================================================
print("\n" + "=" * 70)
print("  ANALYSIS COMPLETE!")
print(f"  {total} columns analyzed")
print(f"  {total*4} plots shown in Spyder Plots panel")
print(f"  {total*4} PNGs saved in 'plots/' folder")
print("  4 Moments CSV : IEX_4Moments_Results.csv")
if os.path.exists(EXCEL_OUTPUT):
    print(f"  Excel output  : {EXCEL_OUTPUT}")
print("=" * 70)